import logging
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app.db import Database

log = logging.getLogger(__name__)


PDF_FONT_CANDIDATES = [
    ("Liberation", "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
     "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf"),
    ("DejaVu", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
     "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"),
]
_pdf_font_registered: str | None = None


def _ensure_pdf_font() -> str:
    global _pdf_font_registered
    if _pdf_font_registered:
        return _pdf_font_registered
    for name, regular, bold in PDF_FONT_CANDIDATES:
        try:
            pdfmetrics.registerFont(TTFont(name, regular))
            pdfmetrics.registerFont(TTFont(f"{name}-Bold", bold))
            _pdf_font_registered = name
            return name
        except Exception as e:
            log.warning("Cannot register %s font: %s", name, e)
    raise RuntimeError("No suitable Cyrillic font found for PDF generation")


@dataclass
class AggregatedRow:
    item_name: str
    unit: str
    supplier: str | None
    cabinet_id: int
    cabinet_name: str
    qty: float
    comments: list[str]
    free_form: bool


async def _aggregate(db: Database, cycle_id: int) -> list[AggregatedRow]:
    """Approved positions, grouped by (catalog_item_id|free_form_name, cabinet)."""
    cur = await db.conn.execute(
        """
        SELECT
            r.id, r.cabinet_id, c.name AS cabinet_name,
            r.catalog_item_id, r.free_form_name,
            r.qty, r.unit, r.comment, r.doctor_name,
            ci.name AS catalog_name, ci.supplier AS catalog_supplier, ci.unit AS catalog_unit
        FROM order_requests r
        JOIN cabinets c ON c.id = r.cabinet_id
        LEFT JOIN catalog_items ci ON ci.id = r.catalog_item_id
        WHERE r.cycle_id = ? AND r.status = 'approved'
        ORDER BY c.id, COALESCE(ci.name, r.free_form_name)
        """,
        (cycle_id,),
    )
    rows = await cur.fetchall()

    grouped: dict[tuple, AggregatedRow] = {}
    for row in rows:
        free_form = row["catalog_item_id"] is None
        name = row["catalog_name"] if not free_form else row["free_form_name"]
        unit = row["catalog_unit"] if not free_form else row["unit"]
        supplier = row["catalog_supplier"] if not free_form else None
        key = (name.lower(), row["cabinet_id"], free_form)

        comment_parts = []
        if row["doctor_name"]:
            comment_parts.append(f"лікар: {row['doctor_name']}")
        if row["comment"]:
            comment_parts.append(row["comment"])
        comment = "; ".join(comment_parts)

        if key in grouped:
            grouped[key].qty += row["qty"]
            if comment:
                grouped[key].comments.append(comment)
        else:
            grouped[key] = AggregatedRow(
                item_name=name,
                unit=unit,
                supplier=supplier,
                cabinet_id=row["cabinet_id"],
                cabinet_name=row["cabinet_name"],
                qty=row["qty"],
                comments=[comment] if comment else [],
                free_form=free_form,
            )
    return list(grouped.values())


def _fmt_qty(qty: float) -> str:
    return f"{qty:g}"


def _summary_text(clinic_name: str, cycle_id: int, opened_at: str, rows: list[AggregatedRow]) -> str:
    if not rows:
        return f"📋 *{clinic_name}*\nЦикл #{cycle_id}: позицій немає."

    by_cabinet: dict[str, list[AggregatedRow]] = defaultdict(list)
    for r in rows:
        by_cabinet[r.cabinet_name].append(r)

    lines = [
        f"📋 *{clinic_name}* — цикл #{cycle_id}",
        f"Відкрито: {opened_at}",
        f"Закрито: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"Всього позицій: *{len(rows)}*",
        "",
    ]
    for cabinet, items in sorted(by_cabinet.items()):
        lines.append(f"*{cabinet}* — {len(items)} поз.")
        for r in items[:15]:
            lines.append(f"  • {r.item_name} — {_fmt_qty(r.qty)} {r.unit}")
        if len(items) > 15:
            lines.append(f"  …і ще {len(items) - 15}")
        lines.append("")
    return "\n".join(lines)


def _build_xlsx(
    clinic_name: str, cycle_id: int, opened_at: str, rows: list[AggregatedRow], path: Path
) -> None:
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0BA373")
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(wrap_text=True, vertical="top")

    def write_sheet(ws, items: list[AggregatedRow], include_cabinet: bool):
        headers = ["№", "Назва"]
        if include_cabinet:
            headers.append("Кабінет")
        headers += ["К-сть", "Од.", "Постачальник", "Коментар", "Free-form"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
        for idx, r in enumerate(items, 1):
            row = [idx, r.item_name]
            if include_cabinet:
                row.append(r.cabinet_name)
            row += [
                r.qty,
                r.unit,
                r.supplier or "",
                "; ".join(r.comments),
                "так" if r.free_form else "",
            ]
            ws.append(row)
        widths = [5, 40]
        if include_cabinet:
            widths.append(14)
        widths += [10, 8, 24, 30, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w
        for col_idx in range(1, len(headers) + 1):
            for cell in list(ws.columns)[col_idx - 1][1:]:
                cell.alignment = wrap

    summary = wb.active
    summary.title = "Зведено"
    write_sheet(summary, rows, include_cabinet=True)

    by_cabinet: dict[str, list[AggregatedRow]] = defaultdict(list)
    for r in rows:
        by_cabinet[r.cabinet_name].append(r)
    for cabinet, items in sorted(by_cabinet.items()):
        ws = wb.create_sheet(cabinet[:31])
        write_sheet(ws, items, include_cabinet=False)

    by_supplier: dict[str, list[AggregatedRow]] = defaultdict(list)
    for r in rows:
        by_supplier[r.supplier or "Інше"].append(r)
    used_titles = {sheet.title for sheet in wb.worksheets}
    for supplier, items in sorted(by_supplier.items()):
        # Excel sheet titles forbid : \ / ? * [ ] and must be unique + ≤31 chars
        safe = "".join(c if c not in r":\/?*[]" else " " for c in supplier)
        title = f"П-{safe}"[:31]
        suffix = 2
        while title in used_titles:
            tail = f" ({suffix})"
            title = f"П-{safe}"[: 31 - len(tail)] + tail
            suffix += 1
        used_titles.add(title)
        ws = wb.create_sheet(title)
        write_sheet(ws, items, include_cabinet=True)

    wb.save(path)


def _build_pdf(
    clinic_name: str, cycle_id: int, opened_at: str, rows: list[AggregatedRow], path: Path
) -> None:
    font = _ensure_pdf_font()
    doc = SimpleDocTemplate(
        str(path),
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
        title=f"Замовлення #{cycle_id}",
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title", parent=styles["Title"], fontName=f"{font}-Bold", fontSize=18, leading=22
    )
    h2 = ParagraphStyle(
        "H2", parent=styles["Heading2"], fontName=f"{font}-Bold", fontSize=13, leading=16
    )
    body = ParagraphStyle("Body", parent=styles["Normal"], fontName=font, fontSize=10, leading=12)

    story = [
        Paragraph(f"{clinic_name}", title_style),
        Paragraph(f"Замовлення витратних матеріалів — цикл #{cycle_id}", body),
        Paragraph(
            f"Відкрито: {opened_at} • Закрито: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            body,
        ),
        Paragraph(f"Всього позицій: {len(rows)}", body),
        Spacer(1, 8 * mm),
    ]

    by_cabinet: dict[str, list[AggregatedRow]] = defaultdict(list)
    for r in rows:
        by_cabinet[r.cabinet_name].append(r)

    def make_table(items: list[AggregatedRow]) -> Table:
        data = [["№", "Назва", "К-сть", "Од.", "Постачальник", "Коментар"]]
        for idx, r in enumerate(items, 1):
            data.append([
                str(idx),
                Paragraph(r.item_name + (" *" if r.free_form else ""), body),
                _fmt_qty(r.qty),
                r.unit,
                r.supplier or "",
                Paragraph("; ".join(r.comments), body),
            ])
        t = Table(data, colWidths=[10 * mm, 70 * mm, 15 * mm, 12 * mm, 32 * mm, 41 * mm], repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0BA373")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), f"{font}-Bold"),
            ("FONTNAME", (0, 1), (-1, -1), font),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#cccccc")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#fafbfc")]),
        ]))
        return t

    for cabinet, items in sorted(by_cabinet.items()):
        story.append(Paragraph(f"{cabinet} — {len(items)} поз.", h2))
        story.append(Spacer(1, 2 * mm))
        story.append(make_table(items))
        story.append(Spacer(1, 6 * mm))

    if any(r.free_form for r in rows):
        story.append(Paragraph("* — позиція додана вручну (не з каталогу)", body))

    doc.build(story)


async def generate_report(
    db: Database,
    cycle_id: int,
    clinic_name: str,
    reports_dir: Path,
) -> tuple[Path, Path, str]:
    cur = await db.conn.execute(
        "SELECT opened_at FROM order_cycles WHERE id = ?", (cycle_id,)
    )
    row = await cur.fetchone()
    opened_at = row["opened_at"] if row else "?"

    rows = await _aggregate(db, cycle_id)
    summary = _summary_text(clinic_name, cycle_id, opened_at, rows)

    out_dir = reports_dir / f"cycle_{cycle_id}"
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M")
    xlsx_path = out_dir / f"order_{cycle_id}_{stamp}.xlsx"
    pdf_path = out_dir / f"order_{cycle_id}_{stamp}.pdf"

    _build_xlsx(clinic_name, cycle_id, opened_at, rows, xlsx_path)
    _build_pdf(clinic_name, cycle_id, opened_at, rows, pdf_path)

    return xlsx_path, pdf_path, summary


async def save_report_record(
    db: Database, cycle_id: int, xlsx: Path, pdf: Path, summary: str
) -> None:
    await db.conn.execute(
        "INSERT INTO reports (cycle_id, xlsx_path, pdf_path, summary_text) VALUES (?, ?, ?, ?)",
        (cycle_id, str(xlsx), str(pdf), summary),
    )
    await db.conn.commit()


async def latest_report(db: Database, cycle_id: int) -> dict | None:
    cur = await db.conn.execute(
        "SELECT xlsx_path, pdf_path, summary_text, generated_at FROM reports "
        "WHERE cycle_id = ? ORDER BY id DESC LIMIT 1",
        (cycle_id,),
    )
    row = await cur.fetchone()
    return dict(row) if row else None
